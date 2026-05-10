from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


_PRIORITY_COLORS = {
    "HIGH": "FFCCCC",
    "MEDIUM": "FFF2CC",
    "LOW": "CCFFCC",
}


def generate_excel(rows: list[tuple[str, str]]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Failed to create Excel worksheet.")
    
    ws.title = "Testing Checklist"

    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")

    ws.append(["Priority", "Test Case"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for priority, test_case in rows:
        ws.append([priority, test_case])
        fill_color = _PRIORITY_COLORS.get(priority)
        if fill_color:
            ws.cell(row=ws.max_row, column=1).fill = PatternFill(fill_type="solid", fgColor=fill_color)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 90

    for row in ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
