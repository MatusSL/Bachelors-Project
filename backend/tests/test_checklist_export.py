from __future__ import annotations

from openpyxl import load_workbook

from backend.core.checklist_parser import parse_checklist
from backend.core.excel_export import generate_excel


def test_parse_checklist_extracts_items_under_priority_headers_only():
    text = """
Intro text is ignored.
[HIGH PRIORITY]
- Verify login rejects expired tokens.
  - Continuation-style markdown is treated as its own item because it starts with a dash.
[MEDIUM PRIORITY] trailing words still match
- Validate pagination does not skip rows.
[LOW PRIORITY]
Not a bullet.
 - Trim whitespace around low priority item.
[UNKNOWN PRIORITY]
- ignored because priority header is unknown
"""

    assert parse_checklist(text) == [
        ("HIGH", "Verify login rejects expired tokens."),
        ("HIGH", "Continuation-style markdown is treated as its own item because it starts with a dash."),
        ("MEDIUM", "Validate pagination does not skip rows."),
        ("LOW", "Trim whitespace around low priority item."),
        ("LOW", "ignored because priority header is unknown"),
    ]


def test_parse_checklist_ignores_bullets_before_first_known_header():
    assert parse_checklist("- orphan item\n[HIGH PRIORITY]\n- included") == [("HIGH", "included")]


def test_generate_excel_creates_workbook_with_headers_rows_widths_and_priority_fills():
    buffer = generate_excel(
        [
            ("HIGH", "High risk item"),
            ("MEDIUM", "Medium risk item"),
            ("LOW", "Low risk item"),
            ("UNKNOWN", "Unknown priority item"),
        ]
    )

    workbook = load_workbook(buffer)
    sheet = workbook.active

    assert sheet.title == "Testing Checklist"
    assert sheet["A1"].value == "Priority"
    assert sheet["B1"].value == "Test Case"
    assert sheet["A1"].font.bold is True
    assert sheet["A1"].alignment.horizontal == "center"
    assert sheet["A2"].value == "HIGH"
    assert sheet["B5"].value == "Unknown priority item"
    assert sheet.column_dimensions["A"].width == 14
    assert sheet.column_dimensions["B"].width == 90
    assert sheet["B2"].alignment.wrap_text is True
    assert sheet["A2"].fill.fgColor.rgb == "00FFCCCC"
    assert sheet["A3"].fill.fgColor.rgb == "00FFF2CC"
    assert sheet["A4"].fill.fgColor.rgb == "00CCFFCC"
    assert sheet["A5"].fill.fill_type is None


def test_generate_excel_handles_empty_rows_with_only_header():
    buffer = generate_excel([])
    sheet = load_workbook(buffer).active

    assert sheet.max_row == 1
    assert sheet.max_column == 2
